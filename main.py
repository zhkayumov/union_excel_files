import os
import openpyxl
import pandas as pd
import datetime
import xlrd
from openpyxl.workbook import Workbook
import pyinputplus as pyip


def open_xls_as_xlsx(filename):
    """
    Данная функция выполняет преобразование .xls файла в .xlsx
    :param filename: Путь до .xls файла, который необходимо преобразовать
    :return:
    """
    try:
        # first open using xlrd
        book = xlrd.open_workbook(filename)
        index = 0
        nrows, ncols = 0, 0
        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1

        # prepare a xlsx sheet
        book1 = Workbook()
        sheet1 = book1.active

        for row in range(1, nrows):
            for col in range(1, ncols):
                sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
        return book1
    except:
        print(f"{filename} Тут расположен очень мутный файл... Как я нек пытался его обработать - не сдюжил ¯\_(ツ)_/¯")
        return None


def concat_excells():
    """
    Основная функция программы.
    Сначала создает список путей до файлов которые нужно обработать и помещает в переменную excell_list. Затем последовательно извлекает нужную информацию из файла и записывает в датафрейм combined.
    В случае, если файл имеет формат .xls - будет осуществлена попытка преобразовать его в .xlsx файл с помощью функции open_xls_as_xlsx().
    В финале работы получившийся датафрейм записывает в эксель файл.
    :return:
    """
    # TODO текущее исполнение не оптимально, т.к. код дублируется для xls и xlsx расширений, можно оптимизировать через функцию или поиграться с расположением параметров if-else.
    excell_list = []
    for root, dirs, files in os.walk("UNZIP"):
        for name in files:
            file_path = os.path.join(root, name)
            excell_list.append(file_path)
    combined = pd.DataFrame()
    for filepath in excell_list:
        filename, file_extension = os.path.splitext(filepath)  # Получение расширения файла
        if file_extension == '.xls':
            filepath2 = filepath.replace(".xls", ".xlsx")
            filez = open_xls_as_xlsx(filepath)
            filez.save(filepath2) if filez != None else filez
            if os.stat(filepath).st_size < 1000000:
                try:
                    wb = openpyxl.load_workbook(filepath2, )
                    ws = wb[wb.sheetnames[0]]  # открываем первый лист в файле
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == 'SAP номер магазина':
                                file = pd.read_excel(filepath2, skiprows=cell.row - 1, usecols=[0, 1, 2, 3])
                                file.columns = ['SAP номер магазина', 'Номер материала', 'Наименование материала',
                                                'Проблема (Дефицит/Перетарка)']
                                task_number = str(os.path.basename(os.path.dirname(filepath2))).replace("allFiles_", '')
                                dann = [f'{task_number}'] * len(file.index)
                                file.insert(4, "Номер заявки", dann, False)
                                combined = pd.concat([combined, file])
                except:
                    task_number = str(os.path.basename(os.path.dirname(filepath2))).replace("allFiles_", '')
                    print(task_number)
            try:
                combined.to_excel("result.xlsx")
            except ValueError:
                combined.to_csv("result.txt")

        elif file_extension == '.xlsx':
            if os.stat(filepath).st_size < 1000000:
                try:
                    wb = openpyxl.load_workbook(filepath, )
                    ws = wb[wb.sheetnames[0]]  # открываем первый лист в файле
                    for row in ws.rows:
                        for cell in row:
                            if cell.value == 'SAP номер магазина':
                                file = pd.read_excel(filepath, skiprows=cell.row - 1, usecols=[0, 1, 2, 3])
                                file.columns = ['SAP номер магазина', 'Номер материала', 'Наименование материала',
                                                'Проблема (Дефицит/Перетарка)']
                                task_number = str(os.path.basename(os.path.dirname(filepath))).replace("allFiles_", '')
                                dann = [f'{task_number}'] * len(file.index)
                                file.insert(4, "Номер заявки", dann, False)
                                combined = pd.concat([combined, file])
                except:
                    task_number = str(os.path.basename(os.path.dirname(filepath))).replace("allFiles_", '')
                    print(task_number)
    try:
        combined.to_excel("result.xlsx")
    except ValueError:
        combined.to_csv("result.txt")


concat_excells()
